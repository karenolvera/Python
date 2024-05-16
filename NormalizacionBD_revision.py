"""
Documento que realiza la normalización de base Rana para el laboratorio de Infantes 
Se inicia con el archivo excel y mediante python se realiza el proceso de obtener los siguientes datos
ID, SEXO, EDAD, PALABRA, TIPO_PALABRA, LOCUTOR 

Karen, mediante CTRL + F en tu IDE 
Buscar "Nota" para checar los comentarios que te he agregado
"""

from openpyxl import workbook
from openpyxl import load_workbook

#Crea la hoja 1
wb2 = load_workbook('5164_30_Limpieza.xlsx')
ws1 = wb2.create_sheet('hoja1')


# Seleccionar la hoja de origen (Definitivo) y la hoja de destino (hoja1)
ws2_origen = wb2['Definitivo']
ws1_destino = wb2['hoja1']

#Columnas que deseas copiar 
"""
NOTA:
 ¿Y si cambian las columnas? (Error humano)
TIP:
 Buscar como referenciar el nombre de la columna
EJEMPLO
    columna_origen_1 = ws2_origen['Locutor']
    columna_origen_1 = ws2_origen['Definitivo']
"""
columna_origen_1 = ws2_origen['H']  
columna_origen_2 = ws2_origen['M']  

# Iterar sobre las celdas en las columnas de origen y copiar los datos a la hoja de destino
for celda_origen_1, celda_origen_2 in zip(columna_origen_1, columna_origen_2):
    """
    #Reducción de lineas de código, 3 lineas en 1
    #¿Se puede hacer esto? 
    ws1_destino.append([celda_origen_1.value, celda_origen_2.value])
    """
    valor_1 = celda_origen_1.value
    valor_2 = celda_origen_2.value
    
    # Agregar los valores a las columnas correspondientes en la hoja de destino
    ws1_destino.append([valor_1, valor_2])
    
    
    # Seleccionar el rango de celdas a partir de B2 en hoja1 y reemplazar "(" por "_"
# for row in ws1_destino.iter_rows(min_row=2, min_col=2, values_only=True):
#     for cell in row:
#         if cell is not None and '(' in cell:
#             ws1_destino[cell.coordinate] = cell.replace('(', '_')
    
    """
    NOTA:
    https://blog.aspose.com/es/cells/split-text-to-column-in-excel-using-python/
    Este tutorial es lo que necesitamos, se puede usar solo es acoplar las variables que tengamos.

    """

  

# Guardar los cambios en el archivo Excel
wb2.save('5164_30_Limpieza.xlsx')
