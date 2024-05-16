from openpyxl import workbook
from openpyxl import load_workbook


#Crea la hoja 1
wb2 = load_workbook('5164_30_Limpieza.xlsx')
ws1 = wb2.create_sheet('hoja1')


# Seleccionar la hoja de origen (Definitivo) y la hoja de destino (hoja1)
ws2_origen = wb2['Definitivo']
ws1_destino = wb2['hoja1']

#Columnas que deseas copiar 
columna_origen_1 = ws2_origen['H']  
columna_origen_2 = ws2_origen['M']  

# Iterar sobre las celdas en las columnas de origen y copiar los datos a la hoja de destino
for celda_origen_1, celda_origen_2 in zip(columna_origen_1, columna_origen_2):
    valor_1 = celda_origen_1.value
    valor_2 = celda_origen_2.value
    
    # Agregar los valores a las columnas correspondientes en la hoja de destino
    ws1_destino.append([valor_1, valor_2])


for row in ws1_destino.iter_rows(min_row=2, min_col=2):
    for cell in row:
        if cell.value is not None and '(' in cell.value:
            # Reemplazar el carácter '(' por '_'
            new_value = cell.value.replace('(', '_')
            # Actualizar el valor de la celda en hoja1
            ws1_destino[cell.coordinate] = new_value


# Procesar la columna B a partir de la fila 2 y dividir los datos utilizando ')' como separador
for row in ws1_destino.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        if cell.value is not None:
            # Dividir los datos en columnas basándonos en el separador ')'
            splitted_values = cell.value.split(')')
            # Insertar los valores divididos en las columnas correspondientes
            for i, val in enumerate(splitted_values):
                ws1_destino.cell(row=cell.row, column=2 + i, value=val)

# Crear una nueva hoja llamada 'hoja2'
ws2 = wb2.create_sheet('hoja2')
# Copiar la columna A de hoja1 a hoja2
for row in ws1_destino.iter_rows(min_row=1, min_col=1, max_col=1):
    for cell in row:
        ws2[cell.coordinate] = cell.value

# Crear la fórmula a aplicar en la celda C1 de la hoja2
formula = '=IF(Definitivo!$C$2=1, 1, 0)'
# Establecer la fórmula en la celda C1 de la hoja2
ws2['C1'] = formula

max_row = ws1_destino.max_row
max_column = ws1_destino.max_column
for row in range(2, max_row + 1):
    for col in range(2, max_column + 1):
        ws2.cell(row=row, column=col).value = f'=IF(hoja1!{ws1_destino.cell(row=row, column=col).coordinate}<>"",IF(hoja2!$A{row}="NI",hoja1!{ws1_destino.cell(row=row, column=col).coordinate}&"_1_"&Definitivo!{ws1_destino.cell(row=row, column=col).coordinate}&"_"&Definitivo!$A$2&"_"&$C$1,hoja1!{ws1_destino.cell(row=row, column=col).coordinate}&"_0_"&Definitivo!$A$2&"_"&$C$1),"")'



# Guardar los cambios en el archivo Excel
wb2.save('5164_30_Limpieza.xlsx')